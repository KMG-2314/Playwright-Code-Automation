import openpyxl

input_path = 'd:/Playwright/Data/Resource_Effort_April-Week-2_Aditya-Input.xlsx'
wb = openpyxl.load_workbook(input_path) # NO data_only so we see formulas
ws = wb['Resource Effort']

print("Input Template Col M (Total Hrs) Formulas:")
for r in range(2, 20):
    val = ws.cell(r, 13).value
    res = ws.cell(r, 5).value
    if res:
        print(f"Row {r} ({res}): {val}")

# Also check Estimation for a resource that might have 0 projection
print("\nChecking Estimation for resources...")
ws_est = wb['Estimation-April']
est_res = set()
for r in range(2, ws_est.max_row + 1):
    name = ws_est.cell(r, 5).value
    if name:
        est_res.add(str(name).strip().title())

print(f"Resources with Estimation entries: {len(est_res)}")
