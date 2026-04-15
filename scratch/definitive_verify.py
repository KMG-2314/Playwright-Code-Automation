import openpyxl

report_path = 'd:/Playwright/Output/Reports/Resource_Effort_Update_W3_20260414_201716.xlsx'
wb = openpyxl.load_workbook(report_path, data_only=False)
ws = wb.active

def check_row(r):
    res = ws.cell(r, 5).value
    proj = ws.cell(r, 2).value
    task = ws.cell(r, 3).value
    w3 = ws.cell(r, 9).value
    return f"Row {r}: Res=[{res}] Proj=[{proj}] Task=[{task}] W3={w3}"

print("Definitive Match Check:")
# Piuesh Jain area was around 26-29
for r in range(25, 35):
    print(check_row(r))

print("\nGrand Totals Check (Row 68):")
for c in range(7, 14):
    print(f"Col {openpyxl.utils.get_column_letter(c)}: {ws.cell(68, c).value}")
