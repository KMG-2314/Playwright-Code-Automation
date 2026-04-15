import openpyxl

input_path = 'd:/Playwright/Data/Quincy-April-Week-2-Input.xlsx'
wb = openpyxl.load_workbook(input_path, data_only=False)
ws = wb.active

print("Searching for Totals...")
for r in range(1, ws.max_row + 1):
    b_val = ws.cell(r, 2).value
    c_val = ws.cell(r, 3).value
    if b_val and 'Total' in str(b_val):
        print(f"B{r}: {b_val}")
    if c_val and 'Total' in str(c_val):
        print(f"C{r}: {c_val}")

# Also find rows with SUM formulas in G-K
for r in range(ws.max_row, 1, -1):
    g_val = ws.cell(r, 7).value
    if g_val and isinstance(g_val, str) and 'SUM' in g_val:
        print(f"Row {r} has formula: {g_val}")
        # Only print a few from bottom
        if r < ws.max_row - 100: break
