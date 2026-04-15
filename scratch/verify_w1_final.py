import openpyxl
from openpyxl.utils import range_boundaries

def get_w1_total(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Resource Effort']
    
    # Map for the last value in merged regions for each column
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = val

    def get_val(r, c):
        if (r, c) in merged_map: return merged_map[(r, c)]
        return ws.cell(r, c).value

    total_w1 = 0.0
    for r in range(2, ws.max_row + 1):
        task = get_val(r, 3)
        if task and str(task).strip().lower() == "grand total":
            break
        w1 = get_val(r, 7)
        if isinstance(w1, (int, float)):
            total_w1 += w1
            
    return total_w1

print(f"Input Total W1: {get_w1_total('d:/Playwright/Data/Resource_Effort_April-Week-2_Aditya-Input.xlsx')}")
print(f"Output Total W1: {get_w1_total('d:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_220428.xlsx')}")
