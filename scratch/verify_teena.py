import openpyxl
path = r"d:\Playwright\Output\Reports\Resource_Effort_April-Week-2_20260412_161533.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

for r in range(2, ws.max_row + 1):
    # Check Project (Col 2), Task (Col 3), Resource (Col 5)
    p = ws.cell(r, 2).value
    t = ws.cell(r, 3).value
    res = ws.cell(r, 5).value
    if res and "Teena" in str(res):
         print(f"Row {r}: Project='{p}', Task='{t}', Res='{res}', M='{ws.cell(r, 13).value}'")
    elif p and "Teena" in str(p):
         print(f"Row {r} (Matched Project): Project='{p}', Task='{t}', Res='{res}', M='{ws.cell(r, 13).value}'")
