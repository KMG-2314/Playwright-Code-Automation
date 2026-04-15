import openpyxl

file = 'Data/Quincy-April-Week-2-Input.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
ws = wb['Resource Effort']

print(f"Max Row: {ws.max_row}")
print("Headers:")
headers = [ws.cell(1, c).value for c in range(1, 15)]
print(headers)

print("\nFirst 5 data rows:")
for r in range(2, 7):
    row_vals = [ws.cell(r, c).value for c in range(1, 15)]
    print(row_vals)

wb_form = openpyxl.load_workbook(file, data_only=False)
ws_form = wb_form['Resource Effort']
print("\nFormula check for Row 2 Col 9 (W3):")
print(ws_form.cell(2, 9).value)
print("Formula check for Row 2 Col 12 (Combined):")
print(ws_form.cell(2, 12).value)
