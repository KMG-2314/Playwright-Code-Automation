import openpyxl
path = r"d:\Playwright\Output\Reports\Resource_Effort_April-Week-2_20260412_161533.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print(f"{'Project':<30} | {'Task':<30} | {'Res':<20}")
for r in range(2, ws.max_row + 1):
    p = ws.cell(r, 2).value
    t = ws.cell(r, 3).value
    res = ws.cell(r, 5).value
    if res and "Rajan" in str(res):
         print(f"{str(p):<30} | {str(t):<30} | {str(res):<20}")
    elif p and "Rajan" in str(p):
         print(f"DEBUG: Project contains Rajan? {p}")
