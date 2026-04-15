import openpyxl

report_path = 'd:/Playwright/Output/Reports/Resource_Effort_Update_W3_20260414_192307.xlsx'
wb = openpyxl.load_workbook(report_path, data_only=False)
ws = wb.active

def check_row(r):
    res = ws.cell(r, 5).value
    proj = ws.cell(r, 2).value
    task = ws.cell(r, 3).value
    w3 = ws.cell(r, 9).value
    total_f = ws.cell(r, 13).value
    border_right = ws.cell(r, 2).border.right.style
    return f"Row {r}: Res=[{res}] Proj=[{proj}] Task=[{task}] W3={w3} M_Form={total_f} Border={border_right}"

print("Checking Vinod (Rows 32-65):")
for r in range(32, 66):
    print(check_row(r))

# Check specific merges
print("\nMerge ranges for Vinod (Row 32-60 area):")
for mr in ws.merged_cells.ranges:
    if mr.min_row >= 32 and mr.max_row <= 70:
        if mr.min_col <= 5 <= mr.max_col:
            print(f"Col E Merge: {mr}")
        if mr.min_col <= 2 <= mr.max_col:
            print(f"Col B Merge: {mr}")
