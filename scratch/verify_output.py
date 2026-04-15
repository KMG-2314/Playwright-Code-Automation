import openpyxl
import os

path = r"d:\Playwright\Output\Reports\Resource_Effort_April-Week-2_20260412_161533.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print(f"{'Row':<4} | {'Resource':<20} | {'Project':<20} | {'Task':<20} | {'M':<5}")
print("-" * 80)

for r in range(2, 50): # Check first 50 rows
    res = ws.cell(r, 5).value
    proj = ws.cell(r, 2).value
    task = ws.cell(r, 3).value
    m = ws.cell(r, 13).value
    print(f"{r:<4} | {str(res):<20} | {str(proj):<20} | {str(task):<20} | {str(m):<5}")
