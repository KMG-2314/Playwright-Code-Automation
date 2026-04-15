import openpyxl

report_path = 'd:/Playwright/Output/Reports/Resource_Effort_Update_W3_20260414_191706.xlsx'
wb = openpyxl.load_workbook(report_path, data_only=False)
ws = wb.active # Should be 'ResourceEffort64v14-1'

print(f"Sheet Name: {ws.title}")

# Check Vinod Ramamurthy
print("\nChecking Vinod Ramamurthy (Row 32-50 area):")
for r in range(32, 60):
    res = ws.cell(r, 5).value
    proj = ws.cell(r, 2).value
    task = ws.cell(r, 3).value
    w3_val = ws.cell(r, 9).value
    total_f = ws.cell(r, 13).value
    
    if res and 'Vinod' in str(res):
        print(f"Row {r}: Res=[{res}] Proj=[{proj}] Task=[{task}] W3={w3_val} TotalForm={total_f}")
    elif not res and task: # Part of a block?
        print(f"Row {r}: (Merged) Proj=[{proj}] Task=[{task}] W3={w3_val} TotalForm={total_f}")

# Check for merges
print("\nMerged ranges containing Col 5 (Resource):")
for mr in ws.merged_cells.ranges:
    if mr.min_col <= 5 <= mr.max_col:
        print(f"Merge: {mr}")
