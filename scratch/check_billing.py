import openpyxl

output_path = 'd:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_222538.xlsx'
wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb['Resource Effort']

print("Resource Check (Target vs Projected):")
last_res = ""
for r in range(2, ws.max_row + 1):
    val = ws.cell(r, 5).value
    if val: last_res = val
    if last_res in ['Rajan Singh', 'Ravindra Singh', 'Vinod Ramamurthy']:
        target = ws.cell(r, 13).value
        comb = ws.cell(r, 12).value
        # If it's the first row of the block, print the name and target
        if val:
            print(f"\nResource: {val} | Target: {target}")
        
        proj_val = sum(ws.cell(r, c).value for c in [9, 10, 11] if ws.cell(r, c).value is not None)
        proj_str = " + ".join(str(ws.cell(r, c).value) for c in [9, 10, 11])
        print(f"  Row {r}: Task=[{ws.cell(r,3).value}] | Proj={proj_str} = {proj_val}")
