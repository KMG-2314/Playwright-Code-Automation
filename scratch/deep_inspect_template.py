import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

def inspect_styles(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    
    print(f"Sheet Name: {ws.title}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    
    print("\nMerged Cells:")
    for merged_range in ws.merged_cells.ranges:
        print(f"  {merged_range}")

    print("\nRow 1-5 details (Potential Headers):")
    for r in range(1, 6):
        row_data = []
        for c in range(1, 15):
            cell = ws.cell(r, c)
            row_data.append(f"{cell.value} (Style: {cell.style})")
        print(f"Row {r}: {row_data}")

    print("\nSample Data Row (Row 10) Styles:")
    for c in range(1, 15):
        cell = ws.cell(10, c)
        print(f"Col {c}: Value={cell.value}")
        print(f"  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}, Color: {cell.font.color.rgb if cell.font.color else 'None'}")
        print(f"  Fill: {cell.fill.patternType}, Color: {cell.fill.start_color.rgb if cell.fill.start_color else 'None'}")
        print(f"  Border: {cell.border}")
        print(f"  Alignment: {cell.alignment.horizontal}, {cell.alignment.vertical}")

if __name__ == "__main__":
    inspect_styles("Data/Quincy-April-Week-2-Input.xlsx")
