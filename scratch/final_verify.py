import openpyxl

report_path = 'd:/Playwright/Output/Reports/Resource_Effort_Update_W3_20260414_192711.xlsx'
wb = openpyxl.load_workbook(report_path, data_only=False)
ws = wb.active

def check_row(r):
    res = ws.cell(r, 5).value
    proj = ws.cell(r, 2).value
    task = ws.cell(r, 3).value
    w3 = ws.cell(r, 9).value
    total_f = ws.cell(r, 13).value
    border_right = ws.cell(r, 2).border.right.style
    return f"Row {r}: Res=[{res}] Proj=[{proj}] Task=[{task}] W3={w3} M_Form={total_f} Border={border_right}"

# Check first few resources
print("Resource Layout Check:")
for r in range(1, 15):
    print(check_row(r))

# Check Vinod area (should be ~32)
print("\nVinod Area Check:")
for r in range(30, 45):
    print(check_row(r))

# Check Merged cells
print("\nMerged cells for Col 5 (Resource):")
for mr in ws.merged_cells.ranges:
    if mr.min_col <= 5 <= mr.max_col:
        print(f"Merge: {mr}")
