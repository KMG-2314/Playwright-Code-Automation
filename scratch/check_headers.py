import openpyxl
path = r"d:\Playwright\Resource_Effort_April-Week-2_Aditya-Input.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active # Usually the first tab is Resource Effort
print("Headers (Row 1):")
print([ws.cell(1, c).value for c in range(1, 15)])
print("Data (Row 2):")
print([ws.cell(2, c).value for c in range(1, 15)])
