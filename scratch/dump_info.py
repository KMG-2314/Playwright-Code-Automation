import openpyxl
import os

input_path = 'd:/Playwright/Data/Resource_Effort_April-Week-2_Aditya-Input.xlsx'
wb = openpyxl.load_workbook(input_path, data_only=True)
ws = wb['Resource Effort']

roles = set()
resources = set()

for r in range(2, ws.max_row + 1):
    role = ws.cell(r, 6).value
    resource = ws.cell(r, 5).value
    if role: roles.add(str(role).strip())
    if resource: resources.add(str(resource).strip())

print("Unique Roles:")
for r in sorted(roles):
    print(f"  - {r}")

print("\nUnique Resources:")
for r in sorted(resources):
    print(f"  - {r}")
