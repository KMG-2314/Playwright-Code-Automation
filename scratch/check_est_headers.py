import openpyxl
path = r"d:\Playwright\Resource_Effort_April-Week-2_Aditya-Input.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
# Find estimation sheet (contains "Estimation")
est_sheet = None
for s in wb.sheetnames:
    if "Estimation" in s:
        est_sheet = s
        break

if not est_sheet:
    print("Estimation sheet not found!")
else:
    ws = wb[est_sheet]
    print(f"Estimation Tab: {est_sheet}")
    print("Headers (Row 1):")
    print([ws.cell(1, c).value for c in range(1, 15)])
    print("Data (Row 2):")
    print([ws.cell(2, c).value for c in range(1, 15)])
