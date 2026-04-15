import openpyxl
from openpyxl.utils import get_column_letter

def check_report(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f"File: {path}")
    print(f"Max Row: {ws.max_row}")
    
    # Check Merges
    merged_cells = list(ws.merged_cells.ranges)
    print(f"Total Merged Ranges: {len(merged_cells)}")
    
    # Check for duplicate project names visible (should be merged)
    last_p = None
    for r in range(2, ws.max_row - 2): # Avoid total rows
        p = ws.cell(r, 2).value
        # If project name is the same as previous and not merged, something is wrong
        # Wait, if it is merged, cell(r, 2) might be None if r > start_row of merge
        pass

    # Find Total Rows
    for r in range(ws.max_row - 5, ws.max_row + 1):
        if r < 1: continue
        val_c = str(ws.cell(r, 3).value or "")
        if "Total" in val_c:
            print(f"Total Row Found at {r}: {val_c}")
            for c in range(7, 12):
                print(f"  Col {get_column_letter(c)}: {ws.cell(r, c).value}")

    # Check for resource names
    resources = set()
    for r in range(2, ws.max_row + 1):
        res = ws.cell(r, 4).value
        if res: resources.add(res)
    print(f"Resources found: {resources}")

check_report("d:/Playwright/Output/Reports/Resource_Effort_Final_W3_20260414_233215.xlsx")
