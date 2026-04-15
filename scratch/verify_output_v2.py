import openpyxl
path = r"d:\Playwright\Output\Reports\Resource_Effort_April-Week-2_20260412_161533.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheetnames:", wb.sheetnames)
ws = wb.active
print("Active sheet:", ws.title)
print("Max row:", ws.max_row)

for r in range(1, 15):
    row_vals = [ws.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")
