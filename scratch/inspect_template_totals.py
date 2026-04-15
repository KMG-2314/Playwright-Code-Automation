import openpyxl

# Using the input path from config/data
input_path = 'd:/Playwright/Data/Quincy-April-Week-2-Input.xlsx'
wb = openpyxl.load_workbook(input_path, data_only=False)
ws = wb.active

print(f"Max Row: {ws.max_row}")

# Check the last 10 rows for Totals
print("\nLast 10 rows of input template:")
for r in range(ws.max_row - 10, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {vals}")

# Check for Project merges in input
print("\nSample Project Merges (Col B):")
for mr in ws.merged_cells.ranges:
    if mr.min_col == 2 and mr.max_col == 2:
        print(f"Project Merge: {mr}")
