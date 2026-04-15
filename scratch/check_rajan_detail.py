import openpyxl

output_path = 'd:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_222538.xlsx'
wb = openpyxl.load_workbook(output_path)
ws = wb['Resource Effort']

print("Rajan Singh Detail:")
last_res = ""
for r in range(2, 30):
    val = ws.cell(r, 5).value
    if val: last_res = val
    if last_res == 'Rajan Singh':
        proj = ws.cell(r, 2).value
        task = ws.cell(r, 3).value
        w1 = ws.cell(r, 7).value
        w3 = ws.cell(r, 9).value
        print(f"Row {r}: [{proj}] {task} | W1={w1} | W3={w3}")
