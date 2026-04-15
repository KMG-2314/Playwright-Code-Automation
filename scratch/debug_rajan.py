import openpyxl

output_path = 'd:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_222538.xlsx'
wb = openpyxl.load_workbook(output_path)
ws = wb['Resource Effort']

print("Rajan Singh Rows Debug:")
for r in range(2, 20):
    res = ws.cell(r, 5).value
    if res == 'Rajan Singh':
        print(f"Row {r}: B={ws.cell(r,2).value} | C={ws.cell(r,3).value} | I={ws.cell(r,9).value} | J={ws.cell(r,10).value} | K={ws.cell(r,11).value}")
