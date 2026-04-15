import openpyxl

file = 'Data/Quincy-April-Week-2-Input.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
ws = wb['ResourceEffort64v14-1']

print("First rows of Vinod:")
last_res = ""
for r in range(2, ws.max_row + 1):
    res = ws.cell(r, 5).value
    if res: last_res = res
    if 'Vinod' in last_res:
        vals = [ws.cell(r, c).value for c in range(1, 15)]
        print(f"Row {r}: {vals}")

wb_form = openpyxl.load_workbook(file, data_only=False)
ws_form = wb_form['ResourceEffort64v14-1']
print("\nFormula check for some rows:")
for r in range(2, 6):
    print(f"Row {r} Col 12 (Combined): {ws_form.cell(r, 12).value}")
