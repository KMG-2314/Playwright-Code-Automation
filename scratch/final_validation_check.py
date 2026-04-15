import openpyxl

report_path = 'd:/Playwright/Output/Reports/Resource_Effort_Update_W3_20260414_201716.xlsx'
wb = openpyxl.load_workbook(report_path, data_only=False)
ws = wb.active

print("Resource Name Positions:")
for r in range(1, 100):
    val = ws.cell(r, 5).value
    if val:
        print(f"Row {r}: {val}")

print("\nTask Names for resource starting at Row 28:")
for r in range(28, 40):
    print(f"Row {r}: {ws.cell(r, 3).value}")
