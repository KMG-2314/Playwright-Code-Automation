import openpyxl

output_path = 'd:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_224740.xlsx'
wb = openpyxl.load_workbook(output_path)
ws = wb['Resource Effort']

print("Rajan Singh Detail (Ultra Clean):")
last_res = ""
for r in range(2, 20):
    val = ws.cell(r, 5).value
    if val: last_res = val
    if last_res == 'Rajan Singh':
        proj = ws.cell(r, 2).value
        task = ws.cell(r, 3).value
        w1 = ws.cell(r, 7).value
        w3 = ws.cell(r, 9).value
        total = ws.cell(r, 12).value
        print(f"Row {r}: [{proj}] {task} | W1={w1} | W3={w3} | L={total}")

print("\nFormula Check (Col M):")
print(f"M2: {ws.cell(2, 13).value}")

# Check for zero rows
zero_rows = 0
for r in range(2, ws.max_row + 1):
    comb = ws.cell(r, 12).value
    # In openpyxl, formula results aren't available in standard load
    # but I can check if all W1..W5 are 0 or None
    row_vals = [ws.cell(r, c).value for c in range(7, 12)]
    if all(v in (0, None, 0.0) for v in row_vals):
        zero_rows += 1
print(f"\nZero-hour rows found: {zero_rows}")
