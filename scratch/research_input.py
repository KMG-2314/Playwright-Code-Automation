import openpyxl

input_path = 'd:/Playwright/Data/Resource_Effort_April-Week-2_Aditya-Input.xlsx'
wb = openpyxl.load_workbook(input_path, data_only=True)

# 1. Resource Order
ws_effort = wb['Resource Effort']
resource_order = []
for r in range(2, ws_effort.max_row + 1):
    res = ws_effort.cell(r, 5).value
    if res:
        res_name = str(res).strip().title()
        if res_name not in resource_order and res_name.lower() != 'grand total':
            resource_order.append(res_name)

print("Resource Order:")
for r in resource_order:
    print(f"  - {r}")

# 2. Rajan Singh's Target
ws_est = wb['Estimation-April']
rajan_target = 0
for r in range(2, ws_est.max_row + 1):
    res = ws_est.cell(r, 5).value
    if res and 'Rajan' in str(res):
        rajan_target = ws_est.cell(r, 13).value
        print(f"Rajan Target on Row {r}: {rajan_target}")
        break
