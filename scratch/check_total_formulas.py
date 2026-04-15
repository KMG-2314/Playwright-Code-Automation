import openpyxl

report_path = 'd:/Playwright/Output/Reports/Resource_Effort_Update_W3_20260414_192711.xlsx'
wb = openpyxl.load_workbook(report_path, data_only=False)
ws = wb.active

print(f"Sheet: {ws.title}, Max Row: {ws.max_row}")

# Find the Total row
total_row = None
for r in range(ws.max_row, 1, -1):
    val = ws.cell(r, 2).value # Check Col B or C for "Total"
    if val and 'Total' in str(val):
        total_row = r
        print(f"Found Total row at {r}: {val}")
        break

if total_row:
    # Check current formulas in Total row
    for c in range(7, 14):
        print(f"Col {c} formula: {ws.cell(total_row, c).value}")

# Check for leading numbers in Tasks
print("\nSample Task Names:")
for r in range(30, 40):
    print(f"Row {r}: {ws.cell(r, 3).value}")
