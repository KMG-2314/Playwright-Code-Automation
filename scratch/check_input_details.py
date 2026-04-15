import openpyxl
from copy import copy

def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)

file = 'Data/Quincy-April-Week-2-Input.xlsx'
wb = openpyxl.load_workbook(file, data_only=False)
ws = wb['ResourceEffort64v14-1']

# Check Vinod again for "missing HRs" potential
print("Vinod's original rows in input:")
for r in range(32, 40):
    val = [ws.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {val}")

# Check borders of row 32
c32 = ws.cell(32, 2)
print(f"\nRow 32 Col B Border: {c32.border}")
