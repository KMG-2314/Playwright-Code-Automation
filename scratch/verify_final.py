import openpyxl
import os

output_path = 'd:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_222538.xlsx'
wb = openpyxl.load_workbook(output_path) # Load with formulas
ws = wb['Resource Effort']

# 1. Resource Order (Column E)
resources = []
for r in range(2, 25): # Just check the top
    v = ws.cell(r, 5).value
    if v and v not in resources:
        resources.append(v)
print("Top Resources in Output:")
for r in resources[:5]:
    print(f"  - {r}")

# 2. Formula Check (Column I - W3)
print("\nFormula Check (W3):")
for r in range(2, 6):
    cell = ws.cell(r, 9)
    print(f"  Row {r}: {cell.value}")

# 3. Target Check (Rajan Singh)
rajan_rows = []
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 5).value == 'Rajan Singh':
        rajan_rows.append(r)

if rajan_rows:
    target = ws.cell(rajan_rows[0], 13).value
    print(f"\nRajan Singh Target (M{rajan_rows[0]}): {target}")
    
# 4. Total Check (W1..W5 sum for Rajan)
wb_data = openpyxl.load_workbook(output_path, data_only=True)
ws_data = wb_data['Resource Effort']
rajan_total = sum(ws_data.cell(r, 12).value for r in rajan_rows) # Combined col L
print(f"Rajan Singh Combined Total (Calculated): {rajan_total}")
