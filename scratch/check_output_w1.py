import openpyxl
import os

output_path = 'd:/Playwright/Output/Reports/Resource_Effort_April-Week-2_20260413_220219.xlsx'
wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb['Resource Effort']

total_w1 = 0.0
resource_w1 = {}

last_res = ""
for r in range(2, ws.max_row + 1):
    res_raw = ws.cell(r, 5).value
    task_raw = ws.cell(r, 3).value
    w1_raw = ws.cell(r, 7).value
    
    if res_raw: last_res = str(res_raw).strip().title()
    
    if task_raw and str(task_raw).strip().lower() == "grand total":
        break
        
    if last_res and isinstance(w1_raw, (int, float)):
        total_w1 += w1_raw
        resource_w1[last_res] = resource_w1.get(last_res, 0) + w1_raw

print(f"Total W1: {total_w1}")
print("By Resource:")
for res, val in sorted(resource_w1.items()):
    print(f"  {res}: {val}")
