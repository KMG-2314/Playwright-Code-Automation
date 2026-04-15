import openpyxl
path = r"d:\Playwright\Resource_Effort_April-Week-2_Aditya-Input.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
est_sheet = [s for s in wb.sheetnames if "Estimation" in s][0]
ws = wb[est_sheet]
# Print row 13 and surrounding
for r in range(10, 20):
    vals = [ws.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {vals}")
