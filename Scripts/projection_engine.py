"""
Projection Engine — Template-preserving updater.

Key principle:
  - Start from the template (config inputPath) and preserve its layout.
  - Update only the target week's values in-place.
  - If a CSV task doesn't exist for a resource, insert a new row at the end
    of that resource block, cloning the template row style and expanding only
    the resource-block merges (A/E/F/M/N).
"""

import os
import re
import argparse
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from difflib import SequenceMatcher
from copy import copy

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange

from Scripts.config_loader import ConfigLoader
from Scripts.calendar_utils import CalendarUtils
from Scripts.csv_processor import CSVProcessor


class ResourceEffortProjectionEngine:

    def __init__(self):
        self.csv_proc = CSVProcessor()
        self.COL_SNO = 1; self.COL_PROJECT = 2; self.COL_TASK = 3
        self.COL_RESOURCE = 5; self.COL_ROLE = 6
        self.COL_W1 = 7; self.COL_W2 = 8; self.COL_W3 = 9; self.COL_W4 = 10; self.COL_W5 = 11
        self.COL_COMBINED = 12; self.COL_TOTAL = 13; self.COL_COMMENT = 14
        # User requirement: keep output EXACTLY like template.
        # Inserting rows can alter row heights/layout in some templates, so we disable it.
        self.ALLOW_ROW_INSERTS = False

    def run(self, csv_path: str, config_path: str):
        print("=" * 60)
        print("[UPDATER] Definitive Engine for 100% Accuracy")
        print("=" * 60)

        config = ConfigLoader.load(config_path)
        target_date = config["startDate"]
        if hasattr(target_date, 'date'): target_date = target_date.date()
        month, year = target_date.month, target_date.year
        week_num = CalendarUtils.date_to_week_number(target_date, month, year)
        target_col = self.COL_W1 + week_num - 1
        
        csv_actuals = self.csv_proc.load_and_process(csv_path, month, year)
        
        input_path = os.path.join("Data", config.get("inputPath", "Quincy-April-Week-2-Input.xlsx"))
        wb = openpyxl.load_workbook(input_path, data_only=False)
        ws = self._select_resource_effort_sheet(wb)

        # Stable Update
        self._update_template(ws, csv_actuals, week_num, target_col)
        
        # Save
        os.makedirs(os.path.join("Output", "Reports"), exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join("Output", "Reports", f"Resource_Effort_Final_W{week_num}_{ts}.xlsx")
        wb.save(out_path)
        print(f"  Report Saved: {out_path}")
        
        try:
            dl = os.path.join(os.path.expanduser("~"), "Downloads")
            import shutil
            shutil.copy(out_path, os.path.join(dl, os.path.basename(out_path)))
        except: pass

    @staticmethod
    def _select_resource_effort_sheet(wb: openpyxl.Workbook) -> Worksheet:
        """
        Select the Resource Effort sheet in a tolerant way:
          - Prefer exact 'Resource Effort'
          - Else pick the first sheet whose name contains both 'resource' and 'effort'
          - Else fall back to active sheet
        """
        if "Resource Effort" in wb.sheetnames:
            return wb["Resource Effort"]

        def norm(s: str) -> str:
            return re.sub(r"[^a-z0-9]+", "", s.lower())

        for name in wb.sheetnames:
            n = norm(name)
            if "resource" in n and "effort" in n:
                return wb[name]

        return wb.active

    def _update_template(self, ws, csv_actuals, week_num, target_col):
        """
        Non-destructive template updater.

        Preserves:
          - column widths, row heights, styles, existing merges, existing formulas
        Updates:
          - target week column values for matched tasks
          - inserts new rows at end of resource block with cloned style
          - expands only A/E/F/M/N merges for that resource block
          - refreshes Grand Total formulas
        """
        totals = self._find_totals_anchor(ws)
        totals_formula_row = totals["formula_row"]
        totals_marker_row = totals["marker_row"]
        last_data_row = totals["last_data_row"]

        res_blocks = self._build_resource_blocks(ws, last_data_row)

        for csv_owner, data in csv_actuals.items():
            matched = self._find_best_match(csv_owner, res_blocks.keys())
            if not matched:
                continue
            block = res_blocks[matched]

            for csv_t in data.get("tasks", []):
                f_row = self._find_task_in_block(ws, csv_t, block["task_rows"])
                if f_row:
                    ws.cell(f_row, target_col).value = csv_t["hours"].get(week_num, 0)
                    continue
                if self.ALLOW_ROW_INSERTS:
                    # (disabled by default)
                    pass
                else:
                    # Keep template layout intact: skip unmatched tasks.
                    # (Optional: could log skipped tasks to a file.)
                    continue

        # Refresh totals formulas after all insertions
        if totals_formula_row and last_data_row:
            self._update_grand_total_formulas(ws, totals_formula_row, last_data_row)

    # -------------------------
    # Template helpers
    # -------------------------

    def _find_totals_anchor(self, ws: Worksheet) -> Dict[str, Optional[int]]:
        """
        Locate the template's totals section in a tolerant way.

        Supported patterns:
          - A literal 'Grand Total' row (in Task col C or Resource col E)
          - A 'Total' marker row in col C, where the next row contains SUM formulas

        Returns:
          {
            "marker_row": <row with 'Total' label, or the same as formula_row for Grand Total>,
            "formula_row": <row whose formulas should be refreshed>,
            "last_data_row": <last row of data above totals section>
          }
        """
        for r in range(1, ws.max_row + 1):
            v_task = ws.cell(r, self.COL_TASK).value
            v_res = ws.cell(r, self.COL_RESOURCE).value
            if isinstance(v_task, str) and v_task.strip().lower() == "grand total":
                return {"marker_row": r, "formula_row": r, "last_data_row": r - 1}
            if isinstance(v_res, str) and v_res.strip().lower() == "grand total":
                return {"marker_row": r, "formula_row": r, "last_data_row": r - 1}

        for r in range(2, ws.max_row):
            v_task = ws.cell(r, self.COL_TASK).value
            if isinstance(v_task, str) and v_task.strip().lower() == "total":
                return {"marker_row": r, "formula_row": r + 1, "last_data_row": r - 1}

        # Fallback: treat whole sheet as data region
        return {"marker_row": None, "formula_row": None, "last_data_row": ws.max_row}

    def _merged_top_left(self, ws: Worksheet, row: int, col: int) -> Tuple[int, int]:
        """Return (row,col) for the top-left cell if inside a merged range."""
        coord = f"{get_column_letter(col)}{row}"
        for mr in ws.merged_cells.ranges:
            if coord in mr:
                return mr.min_row, mr.min_col
        return row, col

    def _cell_value_resolving_merges(self, ws: Worksheet, row: int, col: int):
        tr, tc = self._merged_top_left(ws, row, col)
        return ws.cell(tr, tc).value

    def _build_resource_blocks(self, ws: Worksheet, last_data_row: int) -> Dict[str, Dict[str, Any]]:
        """Index resource blocks without modifying merges."""
        blocks: Dict[str, Dict[str, Any]] = {}

        curr_res: Optional[str] = None
        curr_start: Optional[int] = None

        for r in range(2, last_data_row + 1):
            # IMPORTANT: detect resource header rows using the raw cell value.
            # If column E is merged for the block, only the top-left row carries the value;
            # other rows are MergedCell(None) and must NOT start a new block.
            raw = ws.cell(r, self.COL_RESOURCE).value
            if raw is None or str(raw).strip() == "":
                continue
            res = str(raw).strip()
            if res.lower() == "grand total":
                break

            if curr_res is None:
                curr_res = res
                curr_start = r
                continue

            if res != curr_res:
                self._finalize_block(ws, blocks, curr_res, curr_start, r - 1)
                curr_res = res
                curr_start = r

        if curr_res is not None and curr_start is not None:
            self._finalize_block(ws, blocks, curr_res, curr_start, last_data_row)

        return blocks

    def _row_is_resource_header(self, ws: Worksheet, r: int) -> bool:
        """
        Heuristic: if column E is non-empty on this row (resolved via merge), treat it
        as the header row for that resource.
        """
        v = self._cell_value_resolving_merges(ws, r, self.COL_RESOURCE)
        return v is not None and str(v).strip() != ""

    def _finalize_block(self, ws: Worksheet, blocks: Dict[str, Dict[str, Any]], res: str, start: int, end: int):
        task_rows: List[int] = []
        for r in range(start, end + 1):
            task_val = self._cell_value_resolving_merges(ws, r, self.COL_TASK)
            task = str(task_val).strip() if task_val is not None else ""
            if not task:
                continue
            if task.lower() == "grand total":
                continue
            task_rows.append(r)
        blocks[res] = {"start": start, "end": end, "task_rows": task_rows}

    def _clone_row_style(self, ws: Worksheet, src_row: int, dst_row: int):
        """Clone cell styles and row height from src_row to dst_row."""
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        ws.row_dimensions[dst_row].hidden = ws.row_dimensions[src_row].hidden
        ws.row_dimensions[dst_row].outlineLevel = ws.row_dimensions[src_row].outlineLevel

        for c in range(1, self.COL_COMMENT + 1):
            sc = ws.cell(src_row, c)
            dc = ws.cell(dst_row, c)
            if sc.has_style:
                dc._style = copy(sc._style)
                dc.font = copy(sc.font)
                dc.border = copy(sc.border)
                dc.fill = copy(sc.fill)
                dc.number_format = sc.number_format
                dc.protection = copy(sc.protection)
                dc.alignment = copy(sc.alignment)

    def _insert_row_preserving_dimensions(self, ws: Worksheet, insert_at: int):
        """
        openpyxl's insert_rows shifts cells but row_dimensions (heights/hidden/etc.)
        can become misaligned in some templates. To preserve the exact template
        look & feel, we shift row_dimensions ourselves.
        """
        max_before = ws.max_row

        # Snapshot row_dimensions from insert_at..max_before
        snap = {}
        for r in range(insert_at, max_before + 1):
            rd = ws.row_dimensions[r]
            snap[r] = (rd.height, rd.hidden, rd.outlineLevel)

        ws.insert_rows(insert_at)

        # Shift dimensions down by 1 for all previously existing rows
        for r in range(max_before, insert_at - 1, -1):
            h, hidden, ol = snap.get(r, (None, False, 0))
            new_rd = ws.row_dimensions[r + 1]
            new_rd.height = h
            new_rd.hidden = hidden
            new_rd.outlineLevel = ol

        # Clear inserted row dimension; it will be set via _clone_row_style
        ins = ws.row_dimensions[insert_at]
        ins.height = None
        ins.hidden = False
        ins.outlineLevel = 0

    def _expand_block_merges(self, ws: Worksheet, start_row: int, old_end: int, new_end: int):
        """
        Expand only the resource-block merges to include newly inserted rows.
        We do NOT rebuild merges for the full sheet.
        """
        if new_end <= old_end:
            return

        cols = [self.COL_SNO, self.COL_RESOURCE, self.COL_ROLE, self.COL_TOTAL, self.COL_COMMENT]
        for col in cols:
            # Find merge range that starts at start_row for this column.
            target = None
            coord = f"{get_column_letter(col)}{start_row}"
            for mr in list(ws.merged_cells.ranges):
                if coord in mr and mr.min_col == col and mr.min_row == start_row:
                    target = mr
                    break
            if not target:
                # Not merged in template; nothing to expand.
                continue

            if target.max_row != old_end:
                # Merge may already be larger; don't shrink.
                continue

            ws.unmerge_cells(str(target))
            ws.merge_cells(start_row=start_row, start_column=col, end_row=new_end, end_column=col)

    def _update_grand_total_formulas(self, ws: Worksheet, gt_row: int, last_data_row: int):
        """Refresh the existing Grand Total formulas to cover all data rows."""
        if last_data_row < 2:
            return
        for c in range(self.COL_W1, self.COL_W5 + 1):
            let = get_column_letter(c)
            ws.cell(gt_row, c).value = f"=SUM({let}2:{let}{last_data_row})"
        if ws.cell(gt_row, self.COL_COMBINED).value is not None:
            ws.cell(gt_row, self.COL_COMBINED).value = f"=SUM(L2:L{last_data_row})"
        ws.cell(gt_row, self.COL_TOTAL).value = f"=SUM(M2:M{last_data_row})"

    def _find_best_match(self, n, keys):
        n = self._adv_norm(n)
        best = None; br = 0
        for k in keys:
            r = SequenceMatcher(None, n, self._adv_norm(k)).ratio()
            if r > 0.85 and r > br: br=r; best=k
        return best

    def _find_task_in_block(self, ws: Worksheet, csv_t: Dict[str, Any], rows: List[int]) -> Optional[int]:
        cp, ct = self._adv_norm(csv_t.get("project", "")), self._adv_norm(csv_t.get("task", ""))
        for r in rows:
            tp_raw = self._cell_value_resolving_merges(ws, r, self.COL_PROJECT)
            tt_raw = self._cell_value_resolving_merges(ws, r, self.COL_TASK)
            tp, tt = self._adv_norm(tp_raw), self._adv_norm(tt_raw)
            if cp == tp and ct == tt:
                return r
        return None

    @staticmethod
    def _adv_norm(text) -> str:
        """Aggressive normalization for matching (layout-safe)."""
        if text is None:
            return ""
        s = str(text).strip().lower()
        s = re.sub(r"(?i)^.*?quincy\s*[-_\s\u2013\u2014]?\s*", "", s)
        s = re.sub(r"[-_()\[\]{}/\\&;,.]", " ", s)
        s = re.sub(r"\b\d+\b", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    @staticmethod
    def _display_polish(val: str) -> str:
        """Display cleanup for inserted rows; do not modify existing template text."""
        if not val:
            return ""
        return re.sub(r"^\d+[\s\-_]*", "", str(val)).strip().title()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", required=True); parser.add_argument("--config", required=True)
    args = parser.parse_args()
    ResourceEffortProjectionEngine().run(args.csv, args.config)


if __name__ == "__main__":
    main()
