"""
Estimation Reader — parses the Estimation-{Month} tab.

FIXES:
  1. SIR - stub rows do NOT overwrite last_project, so SIR #EDocs stays
     under its correct parent project (Quincy GoAnywhere for Rajan Singh).
  2. col 13 included in unmerge fill so total_target propagates properly.
  3. First-occurrence guard on total_target per resource.
  4. Project names kept as-is from estimation (no Quincy prefix stripping here).
"""

import re
import openpyxl
from openpyxl.utils import range_boundaries
from typing import Dict, Any

from Scripts.csv_processor import CSVProcessor


class EstimationReader:

    DATA_START_ROW = 2
    COL_PROJECT = 2
    COL_TASK    = 3
    COL_RESOURCE = 5
    COL_ROLE    = 6
    COL_W1      = 7
    COL_TOTAL   = 13

    def read(self, workbook: openpyxl.Workbook, sheet_name: str) -> Dict[str, Any]:
        if sheet_name not in workbook.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found. Checking for fallback...")
            if len(workbook.sheetnames) >= 2:
                sheet_name = workbook.sheetnames[1]
                print(f"  Using tab: '{sheet_name}'")
            else:
                return {}

        ws = workbook[sheet_name]
        self._unmerge_and_fill(ws)
        gt_row = self._find_grand_total_row(ws)

        result = {}
        last_resource = ""
        last_role     = ""
        last_project  = ""
        resource_target_set = set()

        for r in range(self.DATA_START_ROW, gt_row):
            current_resource = self._clean(ws.cell(r, self.COL_RESOURCE).value)
            if current_resource:
                current_resource = current_resource.title()
                if current_resource != last_resource:
                    last_resource = current_resource
                    last_project  = ""
                    last_role     = ""

            resource = last_resource
            if not resource:
                continue

            current_role = self._clean(ws.cell(r, self.COL_ROLE).value)
            if current_role:
                last_role = current_role

            # --- Project: keep as-is, only skip stub rows ---
            raw_project    = ws.cell(r, self.COL_PROJECT).value
            current_project = self._clean_raw(raw_project)
            if current_project and not self._is_stub_project(current_project):
                last_project = current_project

            project = last_project

            # --- Task ---
            raw_task = ws.cell(r, self.COL_TASK).value
            task = CSVProcessor.clean_allocation_name(str(raw_task)) if raw_task else ""

            # --- Weekly hours ---
            hours = {}
            for w in range(1, 6):
                val = ws.cell(r, self.COL_W1 + w - 1).value
                hours[w] = self._to_num(val)
            row_total = sum(hours.values())

            if not task and row_total == 0:
                continue

            total_target_raw = ws.cell(r, self.COL_TOTAL).value
            total_target = self._to_num(total_target_raw)

            if resource not in result:
                result[resource] = {"total_target": 0.0, "role": last_role or "", "tasks": []}

            # Pick the largest target found for this resource (e.g. 161 vs 0)
            if total_target > result[resource]["total_target"]:
                result[resource]["total_target"] = total_target

            if last_role and not result[resource]["role"]:
                result[resource]["role"] = last_role

            if task or row_total > 0:
                # Use fuzzy match for task grouping within the same resource/project
                existing = None
                for t in result[resource]["tasks"]:
                    if t["project"] == project and t["task"] == task:
                        existing = t
                        break
                
                if existing:
                    for w in range(1, 6):
                        existing["hours"][w] += hours[w]
                    existing["total"] += row_total
                else:
                    result[resource]["tasks"].append({
                        "project": project or "",
                        "task":    task    or "",
                        "hours":   hours,
                        "total":   row_total,
                    })

        return result

    def _unmerge_and_fill(self, ws):
        UNMERGE_COLS = {1, 2, 3, 5, 6, 13}
        for mr in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(mr))
            if min_col in UNMERGE_COLS:
                value = ws.cell(min_row, min_col).value
                try:
                    ws.unmerge_cells(str(mr))
                except KeyError:
                    continue
                for r in range(min_row, max_row + 1):
                    ws.cell(r, min_col).value = value

    @staticmethod
    def _is_stub_project(project: str) -> bool:
        """Return True for meaningless stub entries like 'SIR - '."""
        p = re.sub(r"[\s\-_]+", "", project).strip()
        return len(p) <= 4

    @staticmethod
    def _clean_raw(value) -> str:
        if value is None:
            return ""
        s = str(value).strip().replace("\xa0", " ").strip()
        return re.sub(r"\s+", " ", s).strip()

    @staticmethod
    def _clean(value) -> str:
        if value is None:
            return ""
        return str(value).strip().replace("\xa0", " ").strip()

    @staticmethod
    def _to_num(value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip())
        except (ValueError, TypeError):
            return 0.0

    def _find_grand_total_row(self, ws) -> int:
        for row in ws.iter_rows(min_col=1, max_col=6):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "grand total":
                    return cell.row
        last = self.DATA_START_ROW
        for r in range(ws.max_row, self.DATA_START_ROW - 1, -1):
            if any(ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip()
                   for c in [2, 3, 5]):
                last = r + 1
                break
        return last
