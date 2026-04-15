"""
Template W1 Reader — reads frozen W1 actual hours from the Resource Effort tab
of the input template. These are already-submitted hours that must NOT be
overwritten by projection logic.
"""

import openpyxl
import re
from typing import Dict, Any, List, Tuple


class TemplateW1Reader:
    """
    Reads the Resource Effort sheet from the template and returns W1 hours
    per resource/project/task.

    Returns:
    {
        "Rajan Singh": [
            {"project": "Quincy GoAnywhere", "task": "Documentation", "w1": 14.0},
            {"project": "Quincy GoAnywhere", "task": "Development API",  "w1": 10.0},
        ],
        ...
    }
    """

    DATA_START_ROW = 2
    COL_PROJECT    = 2
    COL_TASK       = 3
    COL_RESOURCE   = 5
    COL_W1_START   = 7

    def read(self, workbook: openpyxl.Workbook, actual_weeks: List[int]) -> Tuple[Dict[str, Any], List[str]]:
        sheet_name = "Resource Effort"
        if sheet_name not in workbook.sheetnames:
            print(f"  [WARN] '{sheet_name}' sheet not found in template")
            return {}, []

        ws = workbook[sheet_name]
        self._unmerge_fill(ws)

        result = {}
        resource_order = []
        last_resource = ""
        last_project  = ""
        last_role     = ""

        for r in range(self.DATA_START_ROW, ws.max_row + 1):
            resource_raw = ws.cell(r, self.COL_RESOURCE).value
            project_raw  = ws.cell(r, self.COL_PROJECT).value
            task_raw     = ws.cell(r, self.COL_TASK).value
            role_raw     = ws.cell(r, 6).value

            resource = self._clean(resource_raw)
            if resource:
                resource = resource.title()
                last_resource = resource
                if resource not in resource_order and resource.lower() != "grand total":
                    resource_order.append(resource)

            project  = self._clean(project_raw)
            task     = self._clean(task_raw)
            role     = self._clean(role_raw)

            if (task and task.lower() == "grand total") or (resource and resource.lower() == "grand total"):
                break

            if project:
                last_project = project
            if role:
                last_role = role

            res  = last_resource
            proj = last_project
            rol  = last_role

            if not res or not task:
                continue

            week_hours = {}
            for w in actual_weeks:
                val = ws.cell(r, self.COL_W1_START + w - 1).value
                h = self._to_num(val)
                week_hours[w] = h

            if res not in result:
                result[res] = []

            result[res].append({
                "project": proj,
                "task": task,
                "role": rol,
                "hours": week_hours,
            })

        return result, resource_order

    def _unmerge_fill(self, ws):
        from openpyxl.utils import range_boundaries
        COLS = {2, 3, 5}
        for mr in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(mr))
            if min_col in COLS:
                value = ws.cell(min_row, min_col).value
                try:
                    ws.unmerge_cells(str(mr))
                except KeyError:
                    continue
                for r in range(min_row, max_row + 1):
                    ws.cell(r, min_col).value = value

    @staticmethod
    def _clean(value) -> str:
        if value is None:
            return ""
        s = str(value).strip().replace("\xa0", " ").strip()
        return s if s else ""

    @staticmethod
    def _to_num(value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip())
        except (ValueError, TypeError):
            return 0.0
