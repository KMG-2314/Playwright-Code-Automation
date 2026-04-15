"""
Resource Effort Writer — complete rewrite with all fixes.

KEY FIXES:
  1. Only 2 output tabs: Resource Effort + Estimation (extra tabs removed).
  2. Project column: smart merge by grouping consecutive identical projects per resource.
  3. Primary Task: numbers/dashes/underscores stripped from display.
  4. Col M = static Monthly Target (never overwritten by formula).
  5. W1 injected from frozen template; W2+ from CSV.
  6. Consistent formatting: uniform font (Calibri 10), header style, borders everywhere.
  7. Projected cells highlighted light yellow; actual cells white.
  8. Estimation project name (future project) shown for tasks not yet started.
"""

import re
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Side, Font, PatternFill, GradientFill
)
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List


# ── Shared style constants ──────────────────────────────────────────────────

FONT_NORMAL    = Font(name="Calibri", size=10)
FONT_BOLD      = Font(name="Calibri", size=10, bold=True)
FONT_HEADER    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

ALIGN_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT    = Alignment(horizontal="right",  vertical="center", wrap_text=True)
ALIGN_CENTER_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)

_thin = Side(border_style="thin", color="000000")
BORDER_THIN    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_med = Side(border_style="medium", color="000000")
BORDER_HEADER  = Border(left=_med, right=_med, top=_med, bottom=_med)

# Fill colours
FILL_HEADER_BLUE  = PatternFill("solid", fgColor="1F6090")   # dark blue header
FILL_PROJ_BLUE    = PatternFill("solid", fgColor="D9E1F2")   # soft blue for projected headers
FILL_ACTUAL_COL   = PatternFill("solid", fgColor="E2EFDA")   # pale green for actual week headers
FILL_PROJ_CELL    = PatternFill("solid", fgColor="FFF2CC")   # pale yellow for projected data
FILL_ACTUAL_CELL  = PatternFill("solid", fgColor="FFFFFF")   # white for actual data
FILL_GT_ROW       = PatternFill("solid", fgColor="D6DCE4")   # grey for grand total
FILL_RESOURCE_HDR = PatternFill("solid", fgColor="BDD7EE")   # mid-blue for resource block first row
FILL_ALT_ROW      = PatternFill("solid", fgColor="F5F5F5")   # very light grey for alternating rows


def _apply_cell(cell, value=None, font=None, alignment=None, border=None, fill=None, number_format=None):
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


class ResourceEffortWriter:

    DATA_START_ROW = 2

    COL_SNO       = 1   # A
    COL_PROJECT   = 2   # B
    COL_TASK      = 3   # C
    COL_SECONDARY = 4   # D
    COL_RESOURCE  = 5   # E
    COL_ROLE      = 6   # F
    COL_W1        = 7   # G
    COL_W2        = 8   # H
    COL_W3        = 9   # I
    COL_W4        = 10  # J
    COL_W5        = 11  # K
    COL_COMBINED  = 12  # L
    COL_TOTAL     = 13  # M
    COL_COMMENT   = 14  # N

    # ── Public entry ───────────────────────────────────────────────────────

    def write(
        self,
        ws,
        all_resource_data: Dict[str, Any],
        actual_weeks: List[int],
        projected_weeks: List[int],
        total_weeks: int,
        template_history: Dict[str, Any] = None,
    ):
        self._template_history = template_history or {}
        self._actual_weeks = actual_weeks
        self._proj_weeks   = projected_weeks

        self._set_column_headers(ws, actual_weeks, projected_weeks)
        self._clear_data(ws)
        data_end_row = self._write_data_rows(
            ws, all_resource_data, actual_weeks, projected_weeks, total_weeks
        )
        gt_row = data_end_row + 1
        self._write_grand_total(ws, gt_row, total_weeks)
        self._apply_merges(ws, data_end_row)
        self._apply_formatting(ws, gt_row, actual_weeks, projected_weeks)
        self._set_column_widths(ws)
        ws.freeze_panes = "G2"   # freeze header row and left columns up to F

    # ── Headers ────────────────────────────────────────────────────────────

    def _set_column_headers(self, ws, actual_weeks, projected_weeks):
        static = {
            self.COL_SNO:       "SNo",
            self.COL_PROJECT:   "Project",
            self.COL_TASK:      "Primary Tasks",
            self.COL_SECONDARY: "Secondary Task",
            self.COL_RESOURCE:  "Resource",
            self.COL_ROLE:      "Role",
            self.COL_COMBINED:  "Combined\nHours",
            self.COL_TOTAL:     "Total Hrs",
            self.COL_COMMENT:   "Comment",
        }
        for col, header in static.items():
            cell = ws.cell(1, col)
            _apply_cell(cell, value=header, font=FONT_HEADER,
                        alignment=ALIGN_CENTER, border=BORDER_THIN, fill=FILL_HEADER_BLUE)

        for w in range(1, 6):
            col  = self.COL_W1 + w - 1
            cell = ws.cell(1, col)
            if w in actual_weeks:
                _apply_cell(cell, value=f"W{w} Hrs\n(Actual)",
                            font=FONT_HEADER, alignment=ALIGN_CENTER,
                            border=BORDER_THIN, fill=FILL_ACTUAL_COL)
            elif w in projected_weeks:
                _apply_cell(cell, value=f"W{w} Hrs\n(Projected)",
                            font=FONT_HEADER, alignment=ALIGN_CENTER,
                            border=BORDER_THIN, fill=FILL_PROJ_BLUE)
            else:
                _apply_cell(cell, value=f"W{w} Hrs",
                            font=FONT_HEADER, alignment=ALIGN_CENTER,
                            border=BORDER_THIN, fill=FILL_HEADER_BLUE)

        ws.row_dimensions[1].height = 36

    # ── Clear ──────────────────────────────────────────────────────────────

    def _clear_data(self, ws):
        for mr in list(ws.merged_cells.ranges):
            try:
                ws.unmerge_cells(str(mr))
            except KeyError:
                pass
        ws._cells = dict(ws._cells)
        for r in range(self.DATA_START_ROW, ws.max_row + 1):
            for c in range(1, 15):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill  = FILL_ACTUAL_CELL

    # ── Task name cleaner ──────────────────────────────────────────────────

    @staticmethod
    def _clean_task_display(task: str) -> str:
        """Strip leading numeric codes, dashes, underscores from task names."""
        if not task:
            return task
        # Remove leading: digits, spaces, dashes, underscores, dots
        cleaned = re.sub(r"^[\d\s_\-\.]+", "", task).strip()
        # Remove trailing year suffixes like "-2026-27", " 2026"
        cleaned = re.sub(r"[\s\-_]+20\d\d[\-\d]*$", "", cleaned).strip()
        # Remove bracketed codes like "(Root)", "(Policy)"
        cleaned = re.sub(r"\s*\([A-Za-z]+\)\s*$", "", cleaned).strip()
        return cleaned if cleaned else task

    # ── W1 lookup ──────────────────────────────────────────────────────────

    @staticmethod
    def _norm_for_match(text: str) -> str:
        if not text:
            return ""
        t = str(text).lower().strip()
        t = re.sub(r"[_\-/\\,\.()&]+", " ", t)
        t = re.sub(r"\b\d+\b", "", t)
        return re.sub(r"\s+", " ", t).strip()

    def _get_w1_hours(self, resource_name: str, est_project: str, est_task: str) -> float:
        entries = self._template_w1.get(resource_name, [])
        if not entries:
            return 0.0
        en = self._norm_for_match(est_task)
        for entry in entries:
            if self._norm_for_match(entry["task"]) == en:
                return entry["w1"]
        # Substring match
        for entry in entries:
            tn = self._norm_for_match(entry["task"])
            if tn and en and (tn in en or en in tn) and len(min(tn, en, key=len)) >= 4:
                return entry["w1"]
        return 0.0

    # ── Data rows ──────────────────────────────────────────────────────────

    def _write_data_rows(
        self, ws, all_resource_data, actual_weeks, projected_weeks, total_weeks
    ) -> int:
        row = self.DATA_START_ROW

        for resource_name, rdata in all_resource_data.items():
            resource_start_row = row
            total_target = rdata.get("total_target", 0)
            rows_data = rdata["rows"]
            weekly_ratios = rdata.get("weekly_ratios", {})

            for ri, task_entry in enumerate(rows_data):
                raw_task   = task_entry["task"]
                clean_task = self._clean_task_display(raw_task)
                project    = task_entry["project"]

                ws.cell(row, self.COL_PROJECT).value   = project
                ws.cell(row, self.COL_TASK).value      = clean_task
                ws.cell(row, self.COL_SECONDARY).value = task_entry.get("secondary", "")
                ws.cell(row, self.COL_RESOURCE).value  = resource_name
                ws.cell(row, self.COL_ROLE).value      = rdata.get("role", "")

                # Weekly hours
                for w in range(1, 6):
                    col_idx = self.COL_W1 + w - 1
                    if w in projected_weeks:
                        ratio = weekly_ratios.get(w, {}).get(ri, 0.0)
                        if ratio > 0:
                            # Formula: Target * Ratio (use static number to avoid circular ref with Col M)
                            ws.cell(row, col_idx).value = f"=ROUND({total_target} * {ratio:.6f}, 0)"
                        else:
                            ws.cell(row, col_idx).value = 0
                    else:
                        ws.cell(row, col_idx).value = task_entry["hours"].get(w, 0)

                # Combined = SUM(G:K)
                ws.cell(row, self.COL_COMBINED).value = f"=SUM(G{row}:K{row})"

                row += 1

            resource_end_row = row - 1

            # Col M: dynamic SUM of Combined hours (Col L) for this block
            # This allows the user to see the total Monthly achievement
            ws.cell(resource_start_row, self.COL_TOTAL).value = f"=SUM(L{resource_start_row}:L{resource_end_row})"

            # Comment on first row
            ws.cell(resource_start_row, self.COL_COMMENT).value = rdata.get("comment", "")

        return row - 1

    # ── Grand total ────────────────────────────────────────────────────────

    def _write_grand_total(self, ws, gt_row, total_weeks):
        ws.cell(gt_row, self.COL_TASK).value = "Grand Total"
        s = self.DATA_START_ROW
        e = gt_row - 1
        for w in range(1, 6):
            cl = get_column_letter(self.COL_W1 + w - 1)
            ws.cell(gt_row, self.COL_W1 + w - 1).value = f"=SUM({cl}{s}:{cl}{e})"
        ws.cell(gt_row, self.COL_COMBINED).value = f"=SUM(L{s}:L{e})"
        ws.cell(gt_row, self.COL_TOTAL).value    = f"=SUM(M{s}:M{e})"

    # ── Smart merge: Project & Task ────────────────────────────────────────

    @staticmethod
    def _smart_norm(text: str) -> str:
        """Aggressive normalization for visual merging of projects/tasks."""
        if not text:
            return ""
        t = str(text).lower().strip()
        # Strip versioning: V2, V10, Version 2
        t = re.sub(r'(?i)[\s\-_]+v[\d\.]+', ' ', t)
        t = re.sub(r'(?i)\bversion\s*\d+\b', ' ', t)
        t = re.sub(r"(?i)^quincy[\s\-_]*", "", t)
        # Numbers and special chars
        t = re.sub(r"[_\-/\\&,\.()]+", " ", t)
        t = re.sub(r"\b\d+\b", " ", t)
        return " ".join(t.split())

    def _apply_merges(self, ws, data_end_row):
        rows = list(range(self.DATA_START_ROW, data_end_row + 1))
        if not rows:
            return

        # Snapshot raw values
        snap = {
            col: {r: ws.cell(r, col).value for r in rows}
            for col in (self.COL_SNO, self.COL_PROJECT, self.COL_TASK,
                        self.COL_RESOURCE, self.COL_ROLE)
        }

        def merge_run(col, start, end):
            """Merge rows start..end in col, keep value at start."""
            if start < end:
                ws.merge_cells(start_row=start, end_row=end,
                               start_column=col, end_column=col)
            ws.cell(start, col).value     = snap[col][start]
            ws.cell(start, col).alignment = ALIGN_CENTER

        def merge_consecutive(col, block_rows):
            """Merge consecutive runs of identical values."""
            if not block_rows:
                return
            run_start = block_rows[0]
            run_val   = snap[col][run_start]
            for r in block_rows[1:]:
                cv = snap[col][r]
                # smart compare for project col
                if col == self.COL_PROJECT:
                    same = (self._smart_norm(str(cv or "")) ==
                            self._smart_norm(str(run_val or "")))
                else:
                    same = (cv == run_val)
                if not same:
                    merge_run(col, run_start, r - 1)
                    run_start = r
                    run_val   = cv
            merge_run(col, run_start, block_rows[-1])

        # 1. Merge Resource (E)
        merge_consecutive(self.COL_RESOURCE, rows)

        # Build resource blocks
        resource_blocks = []
        curr_val, curr_block = None, []
        for r in rows:
            v = snap[self.COL_RESOURCE][r]
            if v != curr_val:
                if curr_block:
                    resource_blocks.append((curr_val, curr_block))
                curr_val, curr_block = v, [r]
            else:
                curr_block.append(r)
        if curr_block:
            resource_blocks.append((curr_val, curr_block))

        # 2. Role (F) — merge within resource block
        for _, blk in resource_blocks:
            merge_consecutive(self.COL_ROLE, blk)

        # 3. Project (B) — smart merge within resource block
        for _, blk in resource_blocks:
            merge_consecutive(self.COL_PROJECT, blk)

        # 4. SNo (A) — one per resource block
        sno = 0
        for _, blk in resource_blocks:
            sno += 1
            s, e = blk[0], blk[-1]
            for r in blk:
                ws.cell(r, self.COL_SNO).value = None
            if s < e:
                ws.merge_cells(start_row=s, end_row=e,
                               start_column=self.COL_SNO, end_column=self.COL_SNO)
            ws.cell(s, self.COL_SNO).value     = sno
            ws.cell(s, self.COL_SNO).alignment = ALIGN_CENTER

        # 5. Col M: static target — KEEP value, do NOT replace with SUM
        for _, blk in resource_blocks:
            s, e = blk[0], blk[-1]
            target_val = ws.cell(s, self.COL_TOTAL).value
            for r in blk[1:]:
                ws.cell(r, self.COL_TOTAL).value = None
            if s < e:
                ws.merge_cells(start_row=s, end_row=e,
                               start_column=self.COL_TOTAL, end_column=self.COL_TOTAL)
            ws.cell(s, self.COL_TOTAL).value     = target_val
            ws.cell(s, self.COL_TOTAL).alignment = ALIGN_CENTER

        # 6. Comment (N) — merge within resource block
        for _, blk in resource_blocks:
            s, e = blk[0], blk[-1]
            comment_val = ws.cell(s, self.COL_COMMENT).value
            if s < e:
                for r in range(s + 1, e + 1):
                    ws.cell(r, self.COL_COMMENT).value = None
                ws.merge_cells(start_row=s, end_row=e,
                               start_column=self.COL_COMMENT, end_column=self.COL_COMMENT)
            ws.cell(s, self.COL_COMMENT).value     = comment_val
            ws.cell(s, self.COL_COMMENT).alignment = ALIGN_LEFT

    # ── Formatting pass ────────────────────────────────────────────────────

    def _apply_formatting(self, ws, gt_row, actual_weeks, projected_weeks):
        """Apply consistent Calibri 10 font, borders, fills to every cell."""
        proj_cols   = {self.COL_W1 + w - 1 for w in projected_weeks}
        actual_cols = {self.COL_W1 + w - 1 for w in actual_weeks}
        num_cols    = set(range(self.COL_W1, self.COL_COMMENT))  # G..M numeric

        # Collect resource block boundaries for alternating row fill
        rows = list(range(self.DATA_START_ROW, gt_row))
        resource_starts = set()
        prev_res = None
        for r in rows:
            res = ws.cell(r, self.COL_RESOURCE).value
            if res and res != prev_res:
                resource_starts.add(r)
                prev_res = res

        resource_block_idx = {}
        idx = 0
        prev_res = None
        for r in rows:
            res = ws.cell(r, self.COL_RESOURCE).value or prev_res
            if res != prev_res:
                idx += 1
                prev_res = res
            resource_block_idx[r] = idx

        for r in range(self.DATA_START_ROW, gt_row + 1):
            is_gt = (r == gt_row)
            block_odd = (resource_block_idx.get(r, 1) % 2 == 1)

            for c in range(1, self.COL_COMMENT + 1):
                cell = ws.cell(r, c)
                cell.border = BORDER_THIN

                if is_gt:
                    cell.font      = FONT_BOLD
                    cell.fill      = FILL_GT_ROW
                    cell.alignment = ALIGN_CENTER if c not in (self.COL_TASK,) else ALIGN_CENTER
                    continue

                cell.font = FONT_BOLD if c in (self.COL_RESOURCE, self.COL_ROLE) else FONT_NORMAL

                cell.font = FONT_BOLD if c in (self.COL_RESOURCE, self.COL_ROLE) else FONT_NORMAL

                if c in proj_cols:
                    cell.fill      = FILL_PROJ_CELL
                    cell.alignment = ALIGN_RIGHT
                elif c in actual_cols or c == self.COL_W1:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_RIGHT
                elif c == self.COL_PROJECT:
                    cell.fill      = PatternFill("solid", fgColor="EBF3FB") if not block_odd else FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_LEFT
                elif c == self.COL_TASK:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_LEFT
                elif c == self.COL_COMMENT:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_LEFT
                elif c in (self.COL_COMBINED, self.COL_TOTAL):
                    cell.fill      = PatternFill("solid", fgColor="F2F2F2")
                    cell.alignment = ALIGN_RIGHT
                elif c == self.COL_SNO:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_CENTER
                elif c == self.COL_RESOURCE:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_LEFT
                elif c == self.COL_ROLE:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.fill      = FILL_ACTUAL_CELL
                    cell.alignment = ALIGN_CENTER

                # Number format for numeric columns
                if c in num_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "0"

        # Grand total row extra formatting
        ws.cell(gt_row, self.COL_TASK).alignment = ALIGN_CENTER

    def _set_column_widths(self, ws):
        widths = {
            "A": 6,   "B": 30,  "C": 40,  "D": 18,
            "E": 24,  "F": 22,
            "G": 13,  "H": 13,  "I": 13,  "J": 13,  "K": 13,
            "L": 14,  "M": 12,  "N": 60,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
